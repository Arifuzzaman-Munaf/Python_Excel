# importing necessary libraries

from openpyxl import  workbook,load_workbook

# loading the workBook or Excel file
wb = load_workbook('iris.xlsx')

# loading the active workSheet
ws = wb.active

# to get the value of a cell
print(ws['A1'].value)

# to change the value of a cell
ws['A1'].value = '10000'
wb.save('iris.xlsx')
print(ws['A1'].value)

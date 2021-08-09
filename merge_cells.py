# importing necessary libraries

from openpyxl import  workbook,load_workbook

# loading the workBook or Excel file
wb = load_workbook('Munaf.xlsx')
ws = wb.active

# merging cells
ws.merge_cells("A1:D1")

# unmerging the cells
ws.unmerge_cells("A1:D1")


# saving the progress
wb.save('Munaf.xlsx')

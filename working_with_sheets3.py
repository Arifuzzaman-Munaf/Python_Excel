# importing necessary libraries

from openpyxl import workbook,load_workbook
from openpyxl.utils import get_column_letter

# loading the workBook or Excel file
wb = load_workbook('iris.xlsx')

# printing all the sheets available in iris.xlsx workbook
print(wb.sheetnames)

ws =  wb.active
print(ws)
# retrieving information from all rows and column
for row in range(1,100):
    for col in range(1,5):
        char = get_column_letter(col)
        print(ws[char + str(row)].value)

# importing necessary libraries

from openpyxl import  workbook,load_workbook

# loading the workBook or Excel file
wb = load_workbook('iris.xlsx')

# printing all the sheets available in iris.xlsx workbook
print(wb.sheetnames)


# creating a sheet in iris.xlsx workbook
# added a sheet titled 'Test'
wb.create_sheet('Test')

print(wb.sheetnames)


# saving all progress till now
wb.save('iris.xlsx')